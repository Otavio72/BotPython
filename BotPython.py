import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
import time
import customtkinter as ctk
import threading
import warnings
import shutil
import concurrent.futures
from PIL import Image, ImageTk, ImageSequence


# ==============================
# CONFIGURAÇÕES INICIAIS
# ==============================
caminhoCSV = "CSV"
Atualizar = "Atualizar"
Processados = "Processados"
Erro = os.path.join(Processados, "Erro")

os.makedirs(caminhoCSV, exist_ok=True)
os.makedirs(Atualizar, exist_ok=True)
os.makedirs(Processados, exist_ok=True)
os.makedirs(Erro, exist_ok=True)

bot_rodando = threading.Event()
locks_por_excel = {}

pistas_thread = [
    "mugello", "monza", "nurburgring", "spa", "vallelunga", "brands_hatch","silverstone",
    "laguna_seca"
]


# ==============================
# FUNÇÕES AUXILIARES
# ==============================
def formatar_tempo(voltas):
    minutos = int(voltas // 60)
    segundos = voltas % 60
    return f"{minutos:02d}:{segundos:05.2f}"

def limpar_processados():
    try:
        total = 0
        for pasta in [Processados, Erro]:
            for item in os.listdir(pasta):
                caminho_item = os.path.join(pasta, item)
                if os.path.isfile(caminho_item):
                    os.remove(caminho_item)
                    total += 1

        log(f"🧹 Limpeza concluída. {total} arquivos removidos de 'Processados' e 'Erro'.")
    except Exception as e:
        log(f"❎ Erro ao limpar pastas: {e}")



# ==============================
# FUNÇÕES PRINCIPAIS
# ==============================
def Criar_Excel(caminho_arquivo, relatorios="Relatorios"):
    data = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")

    try:
        dadosCsv = pd.read_csv(caminho_arquivo, sep='\t')

        voltas = dadosCsv.groupby("lapIndex").agg({
            "lap_time": "max",
            "lapNum": "first",
            "carId": "first",
            "trackId": "first",
            "validBin": "first"
        })

        voltas = voltas.sort_values(by='lapIndex')
        voltas.reset_index(drop=True, inplace=True)
        voltas['lap_time'] = voltas['lap_time'].apply(formatar_tempo)
        voltas['validBin'] = voltas['validBin'].map({0: 'Inválida', 1: 'Válida'})
        pista = str(voltas.loc[0, 'trackId'])
        id_virtual = f"{pista}"

        excel_encontrado = None

        for relatorio in os.listdir(relatorios):
            if relatorio.endswith(".xlsx"):
                id_relatorio = relatorio.split(" id-")[-1].replace(".xlsx", "")
                if id_virtual == id_relatorio:
                    excel_encontrado = os.path.join(relatorios, relatorio)
                    break

        if excel_encontrado:
            shutil.move(caminho_arquivo, Atualizar)
            message = f"⚙️ Relatório '{id_virtual}' já existe. CSV movido para pasta Atualizar.\n"
            log(message)
        else:
            voltas.to_excel(f"Relatorios\\{data} id-{id_virtual}.xlsx", index=False, sheet_name='Sheet1')
            arquivo_nome = os.path.basename(caminho_arquivo)
            destino = os.path.join(Processados, arquivo_nome)
            shutil.move(caminho_arquivo, destino)
            message = f"✅ Arquivo Excel criado com sucesso: {data} id-{id_virtual}.xlsx\n"
            log(message)

    except Exception as e:
        arquivo_nome = os.path.basename(caminho_arquivo)
        destino = os.path.join(Erro, arquivo_nome)
        shutil.move(caminho_arquivo, destino)
        message = f"❎ Erro ao processar '{arquivo_nome}': {e}. Arquivo movido para 'Erro'.\n"
        log(message)

def Atualizar_Excel(caminho_arquivo_csv, relatorios="Relatorios"):
    
    try:
        novos_dados = pd.read_csv(caminho_arquivo_csv, sep='\t', low_memory=False)

        voltas = novos_dados.groupby("lapIndex").agg({
            "lap_time": "max",
            "lapNum": "first",
            "carId": "first",
            "trackId": "first",
            "validBin": "first"
        })

        pista = str(voltas.loc[0, 'trackId'])
        id_virtual_CSV = f"{pista}"

        excel_encontrado = None

        for relatorio in os.listdir(relatorios):
            if relatorio.endswith(".xlsx"):
                id_relatorio = relatorio.split(" id-")[-1].replace(".xlsx", "")
                if id_virtual_CSV == id_relatorio:
                    excel_encontrado = os.path.join(relatorios, relatorio)
                    break

        if not excel_encontrado:
            return

        planilhaNome = "Sheet1"
        id = 1

        excel = load_workbook(excel_encontrado)

        while planilhaNome in excel.sheetnames:
            id += 1
            planilhaNome = f"Sheet{id}"

        voltas = voltas.sort_values(by='lapIndex')
        voltas['lap_time'] = voltas['lap_time'].apply(formatar_tempo)
        voltas['validBin'] = voltas['validBin'].map({0: 'Inválida', 1: 'Válida'})

        with pd.ExcelWriter(excel_encontrado, engine='openpyxl', mode='a', if_sheet_exists='overlay') as leitor:
            voltas.to_excel(leitor, sheet_name=planilhaNome, index=False)
            message = f"✅ Dados Atualizados com Sucesso {planilhaNome}, arquivos movidos para Processados\n"
            log(message)

        arquivo_nome = os.path.basename(caminho_arquivo_csv)
        destino = os.path.join(Processados, arquivo_nome)
        shutil.move(caminho_arquivo_csv, destino)


    except Exception as e:
        arquivo_nome = os.path.basename(caminho_arquivo_csv)
        destino = os.path.join(Erro, arquivo_nome)
        shutil.move(caminho_arquivo_csv, destino)
        message = f"❎ Erro ao atualizar '{arquivo_nome}': {e}. Arquivo movido para 'Erro'.\n"
        log(message)



# ==============================
# MONITORAMENTO DE PASTAS
# ==============================
def monitorar_pasta_CSV(caminhoCSV):
    while bot_rodando.is_set():
        arquivos = [os.path.join(caminhoCSV, f) for f in os.listdir(caminhoCSV) if f.endswith(".csv")]
        if arquivos:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                for resultado in executor.map(Criar_Excel, arquivos):
                    pass
        time.sleep(5)


def monitorar_pasta_atualizar(Atualizar):
    while bot_rodando.is_set():
        arquivos_csv = [f for f in os.listdir(Atualizar) if f.endswith(".csv")]

        if not arquivos_csv:
            time.sleep(5)
            continue


        for arquivo in arquivos_csv:
            try:
                caminho_arquivo_csv = os.path.join(Atualizar, arquivo)
                Atualizar_Excel(caminho_arquivo_csv)
            except Exception:
                continue

        # pausa só depois do processamento completo
        time.sleep(5)

    


# ==============================
# CONTROLE DO BOT
# ==============================
def iniciar_bot():
    if not bot_rodando.is_set():
        bot_rodando.set()
        thread_CSV = threading.Thread(target=monitorar_pasta_CSV, args=(caminhoCSV,), daemon=True)
        thread_atualizar = threading.Thread(target=monitorar_pasta_atualizar, args=(Atualizar,), daemon=True)
        thread_CSV.start()
        thread_atualizar.start()
        message = "✅ Iniciado — monitorando pastas...\n"
        iniciar_gif()
        log(message)


def parar_bot():
    if bot_rodando.is_set():
        bot_rodando.clear()
        message = "❎ Bot parado.\n"
        log(message)
        parar_gif()


# ==============================
# INTERFACE GRÁFICA
# ==============================
gui = ctk.CTk()
gui.title("Bot CSV")
gui.geometry("570x430")

tab_container = ctk.CTkTabview(gui)
tab_container.pack(fill="both", expand=True, padx=20, pady=20)

aba1 = tab_container.add("Iniciar")
aba2 = tab_container.add("Status")
aba3 = tab_container.add("Sobre")
aba4 = tab_container.add("Limpesa")

gif = Image.open("gif.gif")
frames = [ImageTk.PhotoImage(frame.copy()) for frame in ImageSequence.Iterator(gif)]
warnings.filterwarnings("ignore", category=UserWarning, module="customtkinter")
frames_index = 0

titulo = ctk.CTkLabel(aba1, text="DATABOT Full-Auto", font=ctk.CTkFont(size=20, weight="bold"))
titulo.pack(pady=10)

label_gif = ctk.CTkLabel(aba1, text="", image=frames[0])
label_gif.pack()

running = False


def atualizar_gif():
    global frames_index
    if running:
        label_gif.configure(image=frames[frames_index])
        frames_index = (frames_index + 1) % len(frames)
    gui.after(50, atualizar_gif)


def iniciar_gif():
    global running
    running = True
    atualizar_gif()


def parar_gif():
    global running
    running = False


terminal_logs = ctk.CTkTextbox(aba2, width=500, height=250)
terminal_logs.pack(pady=20)


def log(message):
    terminal_logs.insert("end", message + "\n")
    terminal_logs.see("end")


btn_iniciar_bot = ctk.CTkButton(aba1, text="Iniciar Bot", command=iniciar_bot)
btn_iniciar_bot.pack(pady=10)

btn_parar_bot = ctk.CTkButton(aba1, text="Parar Bot", command=parar_bot)
btn_parar_bot.pack(pady=5)

sobre_texto = """DATABOT Full-Auto v1.0
Desenvolvido por Otávio

Descrição:
Este bot monitora pastas de arquivos CSV de telemetria de corridas,
gera relatórios Excel organizados por pista, 
atualiza relatórios existentes adicionando novas voltas e mantém 
um histórico de arquivos processados. 

Funcionalidades:
- Modo Full-Automático: CSVs processados automaticamente com logs.
- Logs em tempo real na aba 'Status'.
- Organização automática dos arquivos processados e tratamento de erros.
"""

sobre = ctk.CTkLabel(aba3, text=sobre_texto, justify="left", font=ctk.CTkFont(size=14))
sobre.pack(pady=20, padx=20)


aviso_limpesa_texto = """

Ao apertar o botão "Limpar Processados", 
todos os arquivos na pasta 'Processados' e 'Erro'
serão excluidos.

"""

titulo_limpeza = ctk.CTkLabel(aba4, text="⚠️ Limpeza de Arquivos ⚠️", font=ctk.CTkFont(size=20, weight="bold"))
titulo_limpeza.pack(pady=10)

aviso_limpesa = ctk.CTkLabel(aba4, text=aviso_limpesa_texto, justify="left", font=ctk.CTkFont(size=14))
aviso_limpesa.pack(pady=20)

btn_limpar_processados = ctk.CTkButton(aba4, text="Limpar Processados", command=limpar_processados)
btn_limpar_processados.pack(pady=5)

gui.mainloop()
